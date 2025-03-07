# -*- coding:utf-8 -*-
import os
import shutil
import pandas as pd
from tkinter import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Side, Border, Font
from openpyxl.worksheet.dimensions import RowDimension
from dateutil import parser
import pdfplumber
import numpy as np


class VAS_GUI():
    # 批量获取服务器数据，进行累加操作
    def get_files(self):
        print('数据操作进行中......')
        # 追加的dataFrame的title
        # self.add_data_title = ['用料名称', '品号', '颜色', '规格', '预估单耗']
        self.add_data_title = []
        self.add_data_title_1 = ['品类']
        self.add_data_title_2 = ['颜色', '有效幅宽/规格', '供应商', '用料名称', '单耗', '单位', '缩水率', '损耗率（%）', 'RMB单价', '美金单价', '克重', '成份', '起订量', '小缸费', '生产周期', '备注']
        # 根据勤哲的key匹配对应trimList中的key和value
        self.local_trim_list_file = 'd:\\IZACTrimlist'
        self.trim_list_file_finish = 'd:\\IZACTrimlist结果'
        self.en_cn_file = r'\\192.168.0.3\01-业务一部资料\软件\trimlistToBom\en_cn.xlsx'
        # 删除目录内文件
        if os.path.exists(self.trim_list_file_finish):
            shutil.rmtree(self.trim_list_file_finish)
        os.mkdir(self.trim_list_file_finish)

        # 读取中英文翻译的配置文件
        # 英文行数（默认最大是301行）
        self.max_en = 301
        self.en_cn = {}
        self.get_en_cn_config(self.en_cn_file)

        # 循环本地临时文件，处理合并
        self.table_value = []
        for lroot, ldirs, lfiles in os.walk(self.local_trim_list_file):
            for lfile in lfiles:
                print(lfile)
                # UNDER COLLAR 的值为 SHELL，需要追加一行
                self.underCollarFlag = False
                self.file_to_dataframe(os.path.join(lroot, lfile), str(lfile).replace('.pdf', '').replace('.PDF', ''))

        # 循环读取文件，修改样式
        for root, dirs, files in os.walk(self.trim_list_file_finish):
            for file in files:
                self.change_file_style(os.path.join(root, file))

        # print(self.table_value)
        print('已经完成导出操作！请到D盘【IZACTrimlist结果】中查看文件吧~~~~')
        input('按回车退出 ')

    # 获取中英文对照配置文件
    def get_en_cn_config(self, io):
        # 打开工作表
        wb = load_workbook(io)
        ws = wb.active
        for i in range(1, self.max_en):
            data_en_value = ws.cell(row=i, column=1).value
            data_cn_value = ws.cell(row=i, column=2).value
            self.en_cn[data_en_value] = data_cn_value

    # 修改文件样式
    def change_file_style(self, io):
        # 打开工作表
        wb = load_workbook(io)
        ws = wb.active
        # 调整第列宽
        ws.column_dimensions['A'].width = 27.38
        ws.column_dimensions['B'].width = 25.25
        ws.column_dimensions['C'].width = 25.25
        ws.column_dimensions['D'].width = 25.25
        ws.column_dimensions['E'].width = 25.25

        # 冻结首行
        ws.freeze_panes = 'A3'

        # 定义表头颜色样式为橙色
        header_fill = PatternFill('solid', fgColor='FFFF00')
        header_fill_qing = PatternFill('solid', fgColor='F4BD6C')
        header_fill_lan = PatternFill('solid', fgColor='EE82EE')

        # 定义对齐样式横向居中、纵向居中
        align = Alignment(horizontal='center', vertical='center', wrapText=True)
        # 定义对齐样式纵向居中, 自动换行
        align_center = Alignment(vertical='center', wrapText=True)

        # 定义边样式为细条
        side = Side('thin')
        # 定义边框样式，有底边和右边
        border = Border(top=side, bottom=side, left=side, right=side)
        
        # 设置品号列的样式
        col_dict = {2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13: 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U'}
        align_col = []
        for i in range(2, 21):
            data_value = ws.cell(row=2, column=i).value
            if data_value == '品号':
                ws.column_dimensions[col_dict[i]].width = 38
            elif data_value == '用料名称' or data_value == '备注':
                ws.column_dimensions[col_dict[i]].width = 48
            elif data_value in self.add_data_title_2:
                ws.column_dimensions[col_dict[i]].width = 12
                align_col.append(col_dict[i])
        
        # 品类所在行
        row_use_name = 0
        for i in range(1, self.max_en + 5):
            data_value = ws.cell(row=i, column=1).value
            if data_value == '品类':
                row_use_name = i
                break
        # 设置基本数据的样式
        base_step = 0
        for row in ws.iter_rows(min_row=1, max_row=row_use_name):
            base_step = base_step + 1
            # ws.row_dimensions[base_step].height = 18
            RowDimension(ws, index=base_step, height=True)
            for cell in row:
                cell.alignment = align_center
                cell.font = Font(size=12)
                if str(cell.value) == '款号' or str(cell.value) == 'PO' or str(cell.value) == '款式':
                    # 设置单元格填充颜色
                    cell.fill = header_fill_qing
                    # 设置单元格对齐方式
                    cell.font = Font(bold=True)

        # 设置用料名称单元格格式
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[row_use_name].height = 30
        for cell in ws[row_use_name]:
            # 设置单元格填充颜色
            cell.fill = header_fill
            # 设置单元格对齐方式
            cell.alignment = align
            # 设置单元格边框
            cell.border = border
            cell.font = Font(bold=True)
        step = 0
        # 合并单元格行号
        merge_row_no = 0
        for row in ws.iter_rows(min_row=row_use_name + 1, max_row=ws.max_row):
            step = step + 1
            ws.row_dimensions[row_use_name + step].height = 26
            column_no = 0
            for cell in row:
                column_no = column_no + 1
                # 判断【衬布颜色请按照面料颜色决定】，对应的D列需要背景色为黄色
                if str(cell.value).__contains__('VAS'):
                    # 设置单元格填充颜色
                    cell.fill = header_fill
                    merge_row_no = step + 2
                cell.alignment = align_center
                cell.border = border
        # 合并单元格
        # ws.merge_cells(start_row=merge_row_no, start_column=1, end_row=merge_row_no, end_column=column_no)
        # 居中
        for col in align_col:
            for cell in ws[col]:
                cell.alignment = align

        # 设置打印标题和打印列
        # ws.print_title_rows = '3:1'
        # ws.print_title_cols = "A:E"

        # 设置打印A4纵向
        ws.set_printer_settings(ws.PAPERSIZE_A4, ws.ORIENTATION_PORTRAIT)

        # 所有列设置为一页 逆向思维,先缩放到页面 然后适合高度改为FLASE
        ws.sheet_properties.pageSetUpPr.fitToPage = True  # 此行必须设置
        ws.page_setup.fitToHeight = False

        ws.print_options.gridLines = True  # 页面设置->工作表->网格线

        # 页边距
        ws.page_margins.left = 0.1  # 左
        ws.page_margins.right = 0.1  # 右
        ws.page_margins.top = 0.1  # 上
        ws.page_margins.bottom = 0.1  # 下
        ws.page_margins.header = 0.1  # 页眉
        ws.page_margins.footer = 0.1  # 页脚

        # 保存
        wb.save(io)

    def file_to_dataframe(self, io, lfile):
        pdf = pdfplumber.open(io)
        colorlist= []
        pdf_title = []
        # 打开PDF文件  
        for page in pdf.pages:
            # 提取文本内容 
            text = page.extract_text()
            if not text.__contains__('TRIMS'):
                continue
            # print(page.extract_tables()[1])
            # 提取表格2中非None的颜色列表
            colorlist_tmp = [i for i in page.extract_tables()[1][0] if i != None and i != '']
            # 提取表格2中的数据
            colorlist = colorlist_tmp[1:]
            # 排除颜色列表中PDF含有的关键字
            not_color_key = ['SUPPLIER', 'COMPOSITION', 'DETAILS']
            if len(colorlist) == 0:
                colorlist = [i for i in page.extract_tables()[1][1] if i != None and i != '' and i not in not_color_key]
            pdf_data = page.extract_tables()[1][1:]
            pdf_title = np.append(['品类', '供应商', '备注', '有效幅宽/规格'], colorlist)
            # 有的PDF可能需要修改列的数量，todo
            if len(pdf_data[0]) != len(pdf_title):
                pdf_title = np.append(['品类', '供应商', '备注', '有效幅宽/规格'], colorlist)
            # 去掉数据尾部的空值
            for i in range(len(pdf_data)):
                pdf_data[i] = pdf_data[i][:len(pdf_title)]
            data_df = pd.DataFrame(data=pdf_data, columns=pdf_title)
            data_df.replace('\n', '', regex=True, inplace=True)
            # 检查是否存在“备注”列
            if '备注' in data_df.columns:
                # 如果存在, 就清空
                data_df['备注'] = ''
            break
        # 根据颜色追加列
        self.add_data_title = self.add_data_title_1 + colorlist + self.add_data_title_2
        # 追加品号title
        add_title = []
        for item in colorlist:
            add_title.append('品号')
        table_data = pd.DataFrame(None, columns=self.add_data_title)
        # print(data_df)
        for item in self.add_data_title:
            if item in pdf_title:
                table_data.loc[:, item] = data_df[item]
        # 追加款号行和title行
        title_list = []
        first_row_data = []
        second_row_data = self.add_data_title_1 + add_title + self.add_data_title_2
        for index in range(len(second_row_data)):
            tmp_val = ''
            if index == 0:
                tmp_val = '款号'
            elif index == 1:
                tmp_val = lfile
            elif index == 2:
                tmp_val = 'PO'
            elif index == 3:
                tmp_val = ''
            elif index == 4:
                tmp_val = '款式'
            first_row_data.append(tmp_val)
        title_list.append(first_row_data)
        title_list.append(second_row_data)
        # 首行数据只保留颜色值
        third_row_data = []
        for item in self.add_data_title:
            tmp_item = item
            if item in self.add_data_title_1 or item in self.add_data_title_2:
                tmp_item = ''
            third_row_data.append(tmp_item)
        title_list.append(third_row_data)
        now_row = pd.DataFrame(title_list, columns=self.add_data_title)
        table_data = pd.concat([now_row, table_data]).reset_index(drop=True)
        # 追加空白行及固定行
        blank_row = []
        for i in self.add_data_title:
            blank_row.append('')
        # for i in range(0, 8):
        #     table_data = pd.concat([table_data, pd.DataFrame([blank_row], columns=self.add_data_title)]).reset_index(drop=True)
        # 尾部追加固定内容
        add_data =  pd.DataFrame(None, columns=self.add_data_title)
        # 品类固定列
        m_name = ['兜布', '织带', '有纺衬', '无纺衬', '拉丝衬', '胸兜牌衬', '无纺纸衬', '马鬃', '胸棉', '袖山棉', '袖山鬃', '肩垫', '拉丝直条', '拉丝斜条', '拉丝斜条', '双面胶', '直条', '贴边扦条', '小白带', '加丝中打条', '线色']
        c_code = ['', 'NO.SH6950*顺面料色', 'FW2157*黑/白', 'XH-5050*黑/白', 'XH-NP5050*黑/白', '2346-2HE*黑/白', '254*黑/白', 'FC0179 150 8010*自然色', 'AH-80*黑/白', '688-80*黑/白', 'B409919W*本色', 'BY-Z0328*黑/白', '9332-1*黑/白', '7158*灰/白', '7158*灰/白', '双面胶*透明色', '5850-1*黑/白', 'IS-8330*黑/白', 'YL-C03*黑/白', 'ZD-3030*黑/白', '顺色线']
        vender = ['', '清川', '库夫纳', '金林', '金林', '鑫海', '科德宝', '库夫纳', '佳峰', '佳峰', 'SOCO', '白云', '鑫海', '齐祥', '齐祥', '鑫海', '鑫海', '鑫海', '桥新', '齐祥', '']
        specs = ['', '6mm', '150cm', '100cm', '100cm', '', '100cm', '150cm', '100cm', '100cm', '150cm', '', '1cm', '1.5cm', '2cm', '1cm', '2cm', '1.5cm', '0.3cm', '2cm，1.5cm', '']
        part = ['腰兜附上一层，内部兜袋，前肩条1.2cm宽，前袖笼上端条1.2cm宽，后袖笼条', '内领 领吊', '前片，贴边，领面/领座，大小袖山，后袖笼，胸衬开口', '马面上下，马面开祺1.5CM宽，后下摆，后开祺，袖口，腰兜位，兜盖/兜牙，省尖，胸衬开口衬,领底座,手机兜位，门刀,门襟,后兜牙/后兜口,侧兜口,表兜口,腰面', '前下摆圆，前袖隆弯，前袖窿上端1.5CMX12CM长，后肩衬', '胸兜牌', '所有里兜牙，三角牌', '主胸鬃，主棕封口，次胸鬃，挺肩鬃', '胸衬', '袖山棉条', '袖山鬃', '肩', '止口', '后中缝，侧缝，外袖缝', '后中缝，马面缝一面', '胸兜牌，贴边里面上端大约20CM长', '驳口条', '扦贴边里面，胸衬下层', '袖笼,后领口', '前片马面，后片下摆', '']
        consumption = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
        c_color = ['', '顺面料色', '黑', '黑', '黑', '黑', '黑', '自然', '黑', '黑', '本', '黑', '黑', '灰', '灰', '透明', '黑', '黑', '黑', '黑', '顺色线']
        c_union = ['米', '米', '米', '米', '米', '米', '米', '米', '米', '米', '米', '副', '米', '米', '米', '米', '米', '米', '米', '米', '米']
        # 文件名称的全部变成大写后，第一个字母是P或者包含PANT
        if lfile[:1] == 'P' or 'PANT' in lfile.upper():
            m_name = ['拉链', '裤钩', '兜布', '裤膝', '裤膝', '腰里加工', '无纺衬', '有纺衬', '裤钩垫衬', '拉丝斜条', '板带衬', '线色']
            c_code = ['CFC-39 DS H3 P12*顺面料色', 'HHK-0013','', 'PL-2100-1*黑/白', 'KN7200*顺面料色', '腰里加工', 'XH-5050*黑/白', 'HM050*黑/白', 'EXP155*黑/白', '7158*灰/白', '4947*黑/白', '顺色线']
            vender = ['YKK', '上海柯桥','', '海特恩德', 'YCHT', '', '金林', '恒明', '鑫海', '齐祥', '鑫海', '']
            specs = ['', '', '', '72cm', '72cm', '', '100cm', '150cm', '3cm', '1.5cm', '0.9cm', '']
            part = ['前门襟', '腰头','前后兜袋，裆布，裆布包条，下巾里子，门刀包条', '裤膝', '裤膝', '', '门刀，门襟，后兜牙/后兜口，侧兜口，表兜牙，腰面', '腰面先粘一层无纺衬再粘有纺衬', '裤钩垫衬', '后档', '绊带', '']
            consumption = ['', '', '', '', '', '', '', '', '', '', '', '']
            c_color = ['顺面料色', '','', '黑', '顺面料色', '', '黑', '黑', '黑', '黑', '黑', '顺色线']
            c_union = ['条', '套', '米', '米', '米', '米', '米', '米', '米', '米', '米', '米']
        # vas列添加
        m_name.extend(['VAS', '主标', '洗涤', '条码', '主吊牌（含吊粒）', '可回收涤吊牌', '弹力吊牌', '备扣袋', '塑料袋', '衣架', 'IZAC贴纸', 'IZAC贴纸', '主标', '主吊牌（含吊粒）', '吊牌', '洗涤', '条码洗涤', '备扣袋', '塑料袋', '裤架', '纸箱', '垫板', '胶带', '牛皮纸', '贴纸', '贴纸', '小箱贴', '大箱贴'])
        c_code.extend(['', 'IZAWLN08', 'IZAC缎带印唛', 'IZAC无纺印唛', 'IZAHTN01', 'IZAHPP010-PR', 'IZAHTI02-STR ', 'IZAC备扣袋', 'IZAC塑料袋', 'MAYSHINE 2179', '白底黑色', '白底黑色', 'IZAWLN03', 'IZAHTN02', 'IZAHTI02-JER', 'IZAC缎带印唛', 'IZAC无纺印唛', 'IZAC备扣袋', 'IZAC塑料袋（3%可再生）', 'MAYSHINE 2164', '双层皮筋，无箱唛', '垫板', '6道透明印红字', '80克国产牛皮纸', 'IZAC贴纸', 'IZAC贴纸', 'IZAC贴纸', 'IZAC贴纸'])
        c_color.extend(['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        consumption.extend(['', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1'])
        specs.extend(['', '40*60mm', '50*100mm', '40*95mm', '16.5*3.5CM', '4*3.5CM', '10*5.5cm', '（98，5）*64cm', '46cm', '5.4CM, 10000个/盒', '45mm*35mm', '10*8cm', '25*60mm', '37*75mm', '35*73mm', '16.5*3.5cm', '4*3.5cm', '9.5*6cm', '（82，5）*50cm', '37cm', '80CM*50CM*20CM', '70*40CM', '5.5cm*100Y', '160cm,130m/卷', '45mm*35mm', '10*8cm', '45mm*35mm', '15*15cm'])
        vender.extend(['', '常美', '博美', '博美', '常美', '常美', '常美', '鹏博', '鹏博', '顺淼', '浩源', '浩源', '常美', '常美', '常美', '博美', '博美', '鹏博', '鹏博', '顺淼', '浩源', '浩源', '毅成', '浩源', '浩源', '浩源', '浩源', '浩源'])
        part.extend(['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        c_union.extend(['', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个', '个'])
        add_data.loc[:, '品类'] = m_name
        add_data.loc[:, '有效幅宽/规格'] = specs
        add_data.loc[:, '供应商'] = vender
        add_data.loc[:, '用料名称'] = part
        add_data.loc[:, '单耗'] = consumption
        add_data.loc[:, '颜色'] = c_color
        add_data.loc[:, '单位'] = c_union
        for item in self.add_data_title:
            if item not in self.add_data_title_1 and item not in self.add_data_title_2 and len(item.strip()) != 0:
                add_data.loc[:, item] = c_code
        # 修改"单位"这一列
        # table_data.loc[:2, '单位'] = table_data.loc[:2, '单位']
        table_data.loc[2:,'单位'] = table_data.loc[2:].apply(lambda row: '个' if 'BUTTONS' in str(row['品类']) else '条' if 'ZIPPER' in str(row['品类']) else '个' if 'PINS' in str(row['品类']) else '米', axis=1)
        table_data = pd.concat([table_data, add_data]).reset_index(drop=True)
        # 导出excel
        excelUrl = self.trim_list_file_finish + '\\' + lfile + '.xlsx'
        writer = pd.ExcelWriter(excelUrl, engine='xlsxwriter')
        table_data.to_excel(writer, lfile, index=False, header=None)
        writer.save()
def gui_start():
    VAS = VAS_GUI()
    VAS.get_files()


if __name__ == '__main__':
    gui_start()
