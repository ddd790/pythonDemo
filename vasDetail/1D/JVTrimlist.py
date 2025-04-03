# -*- coding:utf-8 -*-
import os
import shutil
import pandas as pd
from tkinter import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Side, Border, Font
from openpyxl.worksheet.dimensions import RowDimension
from dateutil import parser
import numpy as np


class VAS_GUI():
    # 批量获取服务器数据，进行累加操作
    def get_files(self):
        print('数据操作进行中......')
        # 追加的dataFrame的title
        self.add_data_title_1 = ['Lot', 'ModelCode', 'CATEGORY', 'ModelName']
        self.add_data_title_2 = ['type品类', 'art number品号', 'color颜色', 'cut width/size有效幅宽/规格', 'supplier供应商', 'material details 用料名称：', 'yield单耗', 'unit单位', '损耗率（%）', 'RMB PRICE 单价', '$ PRICE 美金单价', 'weight 克重', 'component成份', 'MOQ起订量', 'surcharge小缸费', 'lead time 生产周期', 'remark备注']
        # 根据勤哲的key匹配对应trimList中的key和value
        self.local_trim_list_file = 'd:\\IZACTrimlist'
        self.trim_list_file_finish = 'd:\\IZACTrimlist结果'
        # 删除目录内文件
        if os.path.exists(self.trim_list_file_finish):
            shutil.rmtree(self.trim_list_file_finish)
        os.mkdir(self.trim_list_file_finish)

        # 读取中英文翻译的配置文件
        # 英文行数（默认最大是301行）
        self.max_en = 301

        # 循环本地临时文件，处理合并
        self.table_value = []
        for lroot, ldirs, lfiles in os.walk(self.local_trim_list_file):
            for lfile in lfiles:
                print(lfile)
                # UNDER COLLAR 的值为 SHELL，需要追加一行
                self.underCollarFlag = False
                self.file_to_dataframe(os.path.join(lroot, lfile), str(lfile).replace('.pdf', '').replace('.PDF', ''))

        # 循环读取文件，修改样式
        for root, dirs, files in os.walk(self.trim_list_file_finish):
            for file in files:
                self.change_file_style(os.path.join(root, file))

        # print(self.table_value)
        print('已经完成导出操作！请到D盘【IZACTrimlist结果】中查看文件吧~~~~')
        input('按回车退出 ')

    # 修改文件样式
    def change_file_style(self, io):
        # 打开工作表
        wb = load_workbook(io)
        # 循环wb文件中的所有sheet
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            # 调整第列宽
            ws.column_dimensions['A'].width = 27.38
            ws.column_dimensions['B'].width = 25.25
            ws.column_dimensions['C'].width = 25.25
            ws.column_dimensions['D'].width = 25.25
            ws.column_dimensions['E'].width = 25.25

            # 冻结首行
            ws.freeze_panes = 'A3'

            # 定义表头颜色样式为橙色
            header_fill = PatternFill('solid', fgColor='FFFF00')
            header_fill_qing = PatternFill('solid', fgColor='F4BD6C')
            header_fill_lan = PatternFill('solid', fgColor='EE82EE')

            # 定义对齐样式横向居中、纵向居中
            align = Alignment(horizontal='center', vertical='center', wrapText=True)
            # 定义对齐样式纵向居中, 自动换行
            align_center = Alignment(vertical='center', wrapText=True)

            # 定义边样式为细条
            side = Side('thin')
            # 定义边框样式，有底边和右边
            border = Border(top=side, bottom=side, left=side, right=side)
            
            # 设置品号列的样式
            col_dict = {2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13: 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R'}
            align_col = []
            for i in range(2, 19):
                data_value = ws.cell(row=2, column=i).value
                if data_value == 'art number品号':
                    ws.column_dimensions[col_dict[i]].width = 38
                elif data_value == 'material details 用料名称：' or data_value == 'remark备注':
                    ws.column_dimensions[col_dict[i]].width = 48
                elif data_value in self.add_data_title_2:
                    ws.column_dimensions[col_dict[i]].width = 12
                    align_col.append(col_dict[i])
            
            # 品类所在行
            row_use_name = 0
            for i in range(1, self.max_en + 5):
                data_value = ws.cell(row=i, column=1).value
                if data_value == 'type品类':
                    row_use_name = i
                    break
            # 设置基本数据的样式
            base_step = 0
            for row in ws.iter_rows(min_row=1, max_row=row_use_name):
                base_step = base_step + 1
                # ws.row_dimensions[base_step].height = 18
                RowDimension(ws, index=base_step, height=True)
                for cell in row:
                    cell.alignment = align_center
                    cell.font = Font(size=12)
                    cell.border = border
                    if str(cell.value) == '款号' or str(cell.value) == 'PO' or str(cell.value) == '款式':
                        # 设置单元格填充颜色
                        cell.fill = header_fill_qing
                        # 设置单元格对齐方式
                        cell.font = Font(bold=True)

            # 设置用料名称单元格格式
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[row_use_name].height = 61.5
            for cell in ws[row_use_name]:
                # 设置单元格填充颜色
                cell.fill = header_fill
                # 设置单元格对齐方式
                cell.alignment = align
                # 设置单元格边框
                cell.border = border
                cell.font = Font(bold=True)
            step = 0
            # 合并单元格行号
            merge_row_no = 0
            for row in ws.iter_rows(min_row=row_use_name + 1, max_row=ws.max_row):
                step = step + 1
                ws.row_dimensions[row_use_name + step].height = 26
                column_no = 0
                for cell in row:
                    column_no = column_no + 1
                    # 判断【衬布颜色请按照面料颜色决定】，对应的D列需要背景色为黄色
                    if str(cell.value).__contains__('VAS') or str(cell.value).__contains__('鬃衬包'):
                        # 设置单元格填充颜色
                        cell.fill = header_fill
                        merge_row_no = step + 2
                    cell.alignment = align_center
                    cell.border = border
            
            # 设置行高
            for i in range(3, 39):
                data_value = ws.cell(row=i, column=1).value
                # 如果data_value包含Fabric，设置行高为135
                if str(data_value).__contains__('Fabric'):
                    ws.row_dimensions[i].height = 135
                elif str(data_value).__contains__('body'):
                    ws.row_dimensions[i].height = 94.5
                elif str(data_value).__contains__('non woven fusible'):
                    ws.row_dimensions[i].height = 108
                elif str(data_value).__contains__('COAT POCKETING*上衣口袋布'):
                    ws.row_dimensions[i].height = 81
                elif str(data_value).__contains__('woven fusible 2'):
                    ws.row_dimensions[i].height = 40.5
            # 合并单元格
            # ws.merge_cells(start_row=merge_row_no, start_column=1, end_row=merge_row_no, end_column=column_no)
            # 居中
            for col in align_col:
                for cell in ws[col]:
                    cell.alignment = align
            
            # 设置A1,C1,E1单元格背景颜色为黄色，字体加粗
            ws['A1'].font = Font(bold=True)
            ws['C1'].font = Font(bold=True)
            ws['E1'].font = Font(bold=True)
            ws['A1'].fill = header_fill
            ws['C1'].fill = header_fill
            ws['E1'].fill = header_fill
            # 设置打印标题和打印列
            # ws.print_title_rows = '3:1'
            # ws.print_title_cols = "A:E"

            # 设置打印A4纵向
            ws.set_printer_settings(ws.PAPERSIZE_A4, ws.ORIENTATION_PORTRAIT)

            # 所有列设置为一页 逆向思维,先缩放到页面 然后适合高度改为FLASE
            ws.sheet_properties.pageSetUpPr.fitToPage = True  # 此行必须设置
            ws.page_setup.fitToHeight = False

            ws.print_options.gridLines = True  # 页面设置->工作表->网格线

            # 页边距
            ws.page_margins.left = 0.1  # 左
            ws.page_margins.right = 0.1  # 右
            ws.page_margins.top = 0.1  # 上
            ws.page_margins.bottom = 0.1  # 下
            ws.page_margins.header = 0.1  # 页眉
            ws.page_margins.footer = 0.1  # 页脚

            # 保存
        wb.save(io)
    def file_to_dataframe(self, io, lfile):
        # 读取Excel文件
        excel_file = pd.ExcelFile(io)
        # 获取所有sheet的名称
        all_sheets = excel_file.sheet_names
        # 过滤掉不需要的sheet名称
        sheets_to_read = [sheet for sheet in all_sheets if sheet not in ['PHOTO SHOOT', 'TRIMS']]
        # 读取剩余sheet的内容
        dataframes = {sheet: excel_file.parse(sheet) for sheet in sheets_to_read}
        # 读取第一个sheet的title
        df_title=pd.read_excel(io, sheet_name=0, header=None).loc[0].values
        tmp_df = pd.DataFrame(columns=df_title)
        # 每个sheet的内容
        for sheet, df in dataframes.items():
            # 重置索引
            df.reset_index(drop=True, inplace=True)
            tmp_df = pd.concat([tmp_df, df], ignore_index=True)
        # 按ModelName分组，并将每个分组存储在字典中
        grouped_data = {name: group for name, group in tmp_df.groupby('Model Name')}
        # 打印每个分组的内容
        for model_name, df in grouped_data.items():
            # 每一个ModelName保存成一个excel文件，文件名为model_name
            self.df_to_file(df, model_name)
    def df_to_file(self, df, file_name):
        with pd.ExcelWriter(self.trim_list_file_finish + '\\' + file_name + '.xlsx') as writer:
            for i, row in df.iterrows():
                table_data = pd.DataFrame(None, columns=self.add_data_title_2)
                # 将每一行转换为 DataFrame 并写入一个新的 sheet
                row_df = pd.DataFrame([row])
                style_number = str(row_df.iloc[0]['Lot #']) + '-' + row_df.iloc[0]['Model Code']
                model_name = row_df.iloc[0]['Model Name']
                category = row_df.iloc[0]['CATEGORY']
                # 保留的列
                main_title = ['Lot #', 'Fabric Art', 'body lining A', 'body lining B', 'Besom/Envelope pocketing lining', 'sleeve lining', 
                              'Button', 'Under collar', 'Pant upper waistband', 'Pant lower waistband', 'Pant pocketing']
                # 删除掉除了main_title的列
                row_df = row_df.drop([col for col in row_df.columns if col not in main_title], axis=1)
                tmp = pd.melt(row_df,id_vars='Lot #',var_name='type品类',value_name='art number品号')
                # 清空品类为Pant piping所在行的内容
                # tmp.loc[tmp['type品类'] == 'Pant piping', 'art number品号'] = ''
                # tmp.loc[tmp['type品类'] == 'Pant piping', 'type品类'] = ''
                # 如果'body lining A', 'body lining B', 'Besom/Envelope pocketing lining'三个列的值相同，将其合并
                if (tmp.loc[tmp['type品类'] == 'body lining A', 'art number品号'].values[0] == tmp.loc[tmp['type品类'] == 'body lining B', 'art number品号'].values[0] 
                    and tmp.loc[tmp['type品类'] == 'body lining A', 'art number品号'].values[0] == tmp.loc[tmp['type品类'] == 'Besom/Envelope pocketing lining', 'art number品号'].values[0]
                    ) or (pd.isna(tmp.loc[tmp['type品类'] == 'body lining A', 'art number品号'].values[0]) 
                          and pd.isna(tmp.loc[tmp['type品类'] == 'body lining B', 'art number品号'].values[0]) 
                          and pd.isna(tmp.loc[tmp['type品类'] == 'Besom/Envelope pocketing lining', 'art number品号'].values[0])):
                    # 将品类为body lining A的行的品类改为body lining A/body lining B/Besom/Envelope pocketing lining
                    tmp.loc[tmp['type品类'] == 'body lining A', 'type品类'] = 'body lining A/body lining B/Besom/Envelope pocketing lining'
                    # 删除品类为body lining B和Besom/Envelope pocketing lining的行
                    tmp = tmp.drop(tmp[tmp['type品类'] == 'body lining B'].index)
                    tmp = tmp.drop(tmp[tmp['type品类'] == 'Besom/Envelope pocketing lining'].index)
                # 追加款号行和title行
                title_list = []
                first_row_data = []
                for index in range(len(self.add_data_title_2)):
                    tmp_val = ''
                    if index == 0:
                        tmp_val = 'STYLE NUMBER*订单号'
                    elif index == 1:
                        tmp_val = ''
                    elif index == 2:
                        tmp_val = 'Model Name*款号'
                    elif index == 3:
                        tmp_val = model_name
                    elif index == 4:
                        tmp_val = 'CATEGORY*type品类'
                    elif index == 5:
                        tmp_val = category
                    first_row_data.append(tmp_val)
                title_list.append(first_row_data)
                title_list.append(self.add_data_title_2)
                # 组装列的数值
                now_row = pd.DataFrame(title_list, columns=self.add_data_title_2)
                tmp_table_data = pd.DataFrame(columns=self.add_data_title_2)
                tmp_table_data['type品类'] = tmp['type品类']
                tmp_table_data['art number品号'] = tmp['art number品号']
                table_data = pd.concat([now_row, tmp_table_data, table_data]).reset_index(drop=True)
                # 复制'type品类'列的内容为Button的行
                button_row = table_data.loc[table_data['type品类'] == 'Button']
                button_row['material details 用料名称：'] = 'sleeve 8,pant 5, spare 2 袖扣8，裤扣5，备扣2'
                # 找到Button的行
                button_row_index = table_data[table_data['type品类'] == 'Button'].index[0]
                # 插入到Button的行的下面
                table_data = pd.concat([table_data.loc[:button_row_index], button_row, table_data.loc[button_row_index + 1:]]).reset_index(drop=True)
                # 插入一行新的内容
                waistband_row = pd.DataFrame({'type品类': ['waistband piping 腰里夹牙'], 'material details 用料名称：': ['waistband piping 腰里夹牙'], 'unit单位': ['M/米']})
                # 找到Pant upper waistband的行
                waistband_row_index = table_data[table_data['type品类'] == 'Pant upper waistband'].index[0]
                # 插入到Pant upper waistband的行的下面
                table_data = pd.concat([table_data.loc[:waistband_row_index], waistband_row, table_data.loc[waistband_row_index + 1:]]).reset_index(drop=True)
                # 'type品类'列的内容为Button的行, material details 用料名称：列的内容为'front 2,spare 1 止口2，备扣1'
                table_data.loc[table_data['type品类'] == 'Button', 'material details 用料名称：'] = 'front 2,spare 1 止口2，备扣1'
                # 'type品类'列的内容包含FABRIC的行, material details 用料名称：列的内容为'前片，马面'
                table_data.loc[table_data['type品类'].str.contains('Fabric'), 'material details 用料名称：'] = 'front,sidebody,back,sleeve,lower flap/besom,welt,welt pocketing one layer,front dart placemat,french facing split,collar,collar stand,pant front/back,waistband,belt,fly,under fly,waistband extension, waistband tab,side pocket facing/patch,back pocket besom/facing 前片，马面，后片，袖，腰兜盖/兜牙，胸兜牌，胸兜布一层，前省布，内台场，领面，领座，裤前/后片，腰面，绊带，前门刀，下门襟，腰探头，腰头小鼻，侧兜垫带/侧兜贴布，后兜牙/垫带'
                table_data.loc[table_data['type品类'].str.contains('Fabric'), 'art number品号'] = str(row_df.iloc[0]['Lot #'])
                table_data.loc[table_data['type品类'].str.contains('body lining'), 'material details 用料名称：'] = 'front lining,sidebody lining,back lining,armshield,inside pocket besom/facing,call besom/facing,triangle tab,inside button loop, flap，lower pocket facing 前里子，马面里子，后里子，汗垫，里兜牙/垫带，手机兜牙/手机兜垫带，三角牌，内扣鼻，兜盖,腰兜垫带'
                table_data.loc[table_data['type品类'].str.contains('sleeve lining'), 'material details 用料名称：'] = 'sleeve,sleeve tape*4 袖里，袖里拉条*4'
                table_data.loc[table_data['type品类'].str.contains('Under collar'), 'material details 用料名称：'] = 'under collar 领底'
                table_data.loc[table_data['type品类'].str.contains('Under collar'), 'color颜色'] = 'FVL-022'
                table_data.loc[table_data['type品类'].str.contains('Under collar'), 'art number品号'] = 'WP3071'
                table_data.loc[table_data['type品类'].str.contains('upper waistband'), 'material details 用料名称：'] = 'upper waistband 腰里上部'
                table_data.loc[table_data['type品类'].str.contains('lower waistband'), 'material details 用料名称：'] = 'lower waistband 腰里下部'
                table_data.loc[table_data['type品类'].str.contains('Pant pocketing'), 'material details 用料名称：'] = 'under fly, front pocketing, back pocketing, cover crotch, pocket bag extension 下巾里，前兜布，后兜布，裆布，侧兜拉布'
                # 追加空白行及固定行
                blank_row = []
                for i in self.add_data_title_2:
                    blank_row.append('')
                # 尾部追加固定内容
                add_data =  pd.DataFrame(None, columns=self.add_data_title_2)
                # 向add_data中追加数据
                add_data.loc[len(add_data.index)] = ['pin 胸针','','','','','lapel buttnhole 驳头扣眼','1','PC/个','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['COAT POCKETING*上衣口袋布','K10110-4','','','','inside pocketing, welt pocketing one layer, back neck tape,front shouler tape 45°，front armhole tape 45°,back armhole tape,straight tape,open seam tape,lower pocketing/tape，chest tape里兜布，胸兜布一层，后领口布，前肩条，前袖窿条，后袖窿条，直条，肩劈缝条，腰兜布/拉条,胸衬拉条','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['ZIPPER*拉链','','','','','front fly 前门刀','','PC/条','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['KNEE LINING*裤膝','','','','','front half knee lining 前片半裤膝','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['鬃衬包','','','','','','','','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['woven fusible 1 有纺衬1','BVM50','白','148cm','科德宝','front,lower flap,前片衬，腰兜盖衬','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['woven fusible 2 有纺衬2','ME9013','白','148cm','科德宝','facing,lower pocket position, collar point，inside facing split pocket position 贴边衬,腰兜位衬，领头衬，里兜位','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['main canvas 主鬃','K380C','自然','158cm','科德宝','main canvas，canvas placemat 主鬃，胸衬毛棕垫','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['second canvas 次鬃','K380C','自然','158cm','科德宝','second canvas 次鬃','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['shoulder canvas 肩鬃','KO183','自然','55cm','科德宝','shoulder canvas 肩鬃','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['chest felt 胸棉','KW997','白','145cm','科德宝','chest felt 胸棉','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['sleeve canvas 袖山鬃','KO361C','自然','158cm','科德宝','sleeve head canvas front，back，canvas tape，袖山鬃，前，后，毛棕小条X1，毛棕大条X1','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['sleeve felt 袖山棉条','GO414','黑','158cm','科德宝','sleeve head felt 袖山棉条','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['shoulder pad 肩垫','','黑','','','shoulder 肩','1','PC/付','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['non woven fusible 无纺衬','PE125','灰','150cm','科德宝','sleeve cuff,back vent,back bottom,back neck,sidebody bottom,front neck,facing neck,french facing,front dart point,pant back pocket besom/ position,side pocket opening,fly,under fly,waistband extension/tab 袖口衬，后开祺衬，后下摆衬，后领口衬，马面下摆衬，前片领口，贴边领口，台场衬，省尖衬，裤子后兜牙/位，侧斗口，上巾，下巾，腰探头，腰头小鼻','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['under collar fusible 领底衬','735-4000 ','白','107m','科德宝','under collar 领底','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['tape 无胶衬','SF-35 ','白','99cm','鑫海','inside besom, lower pocket besom 里兜牙，腰兜牙','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['tape 胸兜牌衬','BW70','白','90cm','科德宝','welt 胸兜牌','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['tape 拉丝无纺衬条 ','9332-1 ','白','1.0cm','鑫海','front edge 前止口','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['tape 双面胶 上衣+裤子','双面胶 ','白','0.8cm','鑫海','coats,pants 上衣，裤子','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['tape 小棉带','小棉带9950-1','白','0.3cm','鑫海','collar felt, front/side body armhole，flap 领绒聚量，前片，马面袖窿，兜盖聚量','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['tape 贴边扦条','3300','白','1.5cm','鑫海','扦贴边','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['tape 绊带衬','4947','白','0.9cm','鑫海','belt loop 绊带','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['tape 腰网衬','6148','白','5.5cm','鑫海','inside waistband  腰里','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['tape 腰硬衬','S300P','白','3.3cm','清川','waistband 腰面','','M/米','','','','','','','','','']
                add_data.loc[len(add_data.index)] = ['tape 线色','','','','','','','M/米','','','','','','','','','']
                # 循环添加30个空白行
                for i in range(30):
                    add_data.loc[len(add_data.index)] = ['','','','','','','','','','','','','','','','','']
                # 添加VAS的行
                add_data.loc[len(add_data.index)] = ['VAS','','','','','','','','','','','','','','','','']
                # 循环添加11个空白行
                for i in range(11):
                    add_data.loc[len(add_data.index)] = ['','','','','','','','PC/个','','','','','','','','','']
                # 修改"单位"这一列
                table_data.loc[2:,'unit单位'] = table_data.loc[2:].apply(lambda row: 'PC/个' if ('BUTTONS' in str(row['type品类']) or 'Button' in str(row['type品类'])) else 'PC/条' if 'ZIPPER' in str(row['type品类']) else 'PC/个' if 'PINS' in str(row['type品类']) else 'M/米', axis=1)
                table_data = pd.concat([table_data, add_data]).reset_index(drop=True)
                # 导出excel
                table_data.to_excel(writer, sheet_name=str(row_df.iloc[0]['Lot #']), index=False, header=None)
def gui_start():
    VAS = VAS_GUI()
    VAS.get_files()

if __name__ == '__main__':
    gui_start()
