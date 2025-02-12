# -*- coding:utf-8 -*-
import os
import shutil
import pandas as pd
from tkinter import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Side, Border, Font
from openpyxl.worksheet.dimensions import RowDimension
from dateutil import parser
import pdfplumber
import numpy as np


class VAS_GUI():
    # 批量获取服务器数据，进行累加操作
    def get_files(self):
        print('数据操作进行中......')
        # 追加的dataFrame的title
        self.add_data_title = []
        self.add_data_title_1 = ['款号', '颜色']
        self.add_data_title_2 = []
        # 根据勤哲的key匹配对应trimList中的key和value
        self.local_trim_list_file = 'd:\\IZACTrimlist'
        self.trim_list_file_finish = 'd:\\IZACTrimlist结果'
        # 删除目录内文件
        if os.path.exists(self.trim_list_file_finish):
            shutil.rmtree(self.trim_list_file_finish)
        os.mkdir(self.trim_list_file_finish)

        # 循环本地临时文件，处理合并
        self.table_value = []
        self.error_file = []
        for lroot, ldirs, lfiles in os.walk(self.local_trim_list_file):
            for lfile in lfiles:
                try:
                    # UNDER COLLAR 的值为 SHELL，需要追加一行
                    self.underCollarFlag = False
                    self.file_to_dataframe(os.path.join(lroot, lfile), str(lfile).replace('.pdf', '').replace('.PDF', ''))
                except:
                    self.error_file.append(lfile)

        # 循环读取文件，合并成一个excel文件
        self.merge_excel_files()
        # 循环读取文件，修改样式
        # for root, dirs, files in os.walk(self.trim_list_file_finish):
        #     for file in files:
        #         self.change_file_style(os.path.join(root, file))

        # print(self.table_value)
        print('已经完成导出操作！请到D盘【IZACTrimlist结果】中查看文件吧~~~~')
        # 打印输出self.error_file
        print('错误文件：' + ','.join(self.error_file))
        input('按回车退出 ')

    # 修改文件样式
    def change_file_style(self, io):
        # 打开工作表
        wb = load_workbook(io)
        ws = wb.active
        # 调整第列宽
        ws.column_dimensions['A'].width = 27.38
        ws.column_dimensions['B'].width = 25.25
        ws.column_dimensions['C'].width = 25.25
        ws.column_dimensions['D'].width = 25.25
        ws.column_dimensions['E'].width = 25.25

        # 冻结首行
        ws.freeze_panes = 'A3'

        # 定义表头颜色样式为橙色
        header_fill = PatternFill('solid', fgColor='FFFF00')
        header_fill_qing = PatternFill('solid', fgColor='F4BD6C')
        header_fill_lan = PatternFill('solid', fgColor='EE82EE')

        # 定义对齐样式横向居中、纵向居中
        align = Alignment(horizontal='center', vertical='center', wrapText=True)
        # 定义对齐样式纵向居中, 自动换行
        align_center = Alignment(vertical='center', wrapText=True)

        # 定义边样式为细条
        side = Side('thin')
        # 定义边框样式，有底边和右边
        border = Border(top=side, bottom=side, left=side, right=side)
        
        # 设置品号列的样式
        col_dict = {2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13: 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U'}
        align_col = []
        for i in range(2, 21):
            data_value = ws.cell(row=2, column=i).value
            if data_value == '品号':
                ws.column_dimensions[col_dict[i]].width = 38
            elif data_value == '用料名称' or data_value == '备注':
                ws.column_dimensions[col_dict[i]].width = 48
            elif data_value in self.add_data_title_2:
                ws.column_dimensions[col_dict[i]].width = 12
                align_col.append(col_dict[i])
        
        # 品类所在行
        row_use_name = 0
        for i in range(1, self.max_en + 5):
            data_value = ws.cell(row=i, column=1).value
            if data_value == '品类':
                row_use_name = i
                break
        # 设置基本数据的样式
        base_step = 0
        for row in ws.iter_rows(min_row=1, max_row=row_use_name):
            base_step = base_step + 1
            # ws.row_dimensions[base_step].height = 18
            RowDimension(ws, index=base_step, height=True)
            for cell in row:
                cell.alignment = align_center
                cell.font = Font(size=12)
                if str(cell.value) == '款号' or str(cell.value) == 'PO' or str(cell.value) == '款式':
                    # 设置单元格填充颜色
                    cell.fill = header_fill_qing
                    # 设置单元格对齐方式
                    cell.font = Font(bold=True)

        # 设置用料名称单元格格式
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[row_use_name].height = 30
        for cell in ws[row_use_name]:
            # 设置单元格填充颜色
            cell.fill = header_fill
            # 设置单元格对齐方式
            cell.alignment = align
            # 设置单元格边框
            cell.border = border
            cell.font = Font(bold=True)
        step = 0
        # 合并单元格行号
        merge_row_no = 0
        for row in ws.iter_rows(min_row=row_use_name + 1, max_row=ws.max_row):
            step = step + 1
            ws.row_dimensions[row_use_name + step].height = 26
            column_no = 0
            for cell in row:
                column_no = column_no + 1
                # 判断【衬布颜色请按照面料颜色决定】，对应的D列需要背景色为黄色
                if str(cell.value).__contains__('VAS'):
                    # 设置单元格填充颜色
                    cell.fill = header_fill
                    merge_row_no = step + 2
                cell.alignment = align_center
                cell.border = border
        # 合并单元格
        # ws.merge_cells(start_row=merge_row_no, start_column=1, end_row=merge_row_no, end_column=column_no)
        # 居中
        for col in align_col:
            for cell in ws[col]:
                cell.alignment = align

        # 设置打印标题和打印列
        # ws.print_title_rows = '3:1'
        # ws.print_title_cols = "A:E"

        # 设置打印A4纵向
        ws.set_printer_settings(ws.PAPERSIZE_A4, ws.ORIENTATION_PORTRAIT)

        # 所有列设置为一页 逆向思维,先缩放到页面 然后适合高度改为FLASE
        ws.sheet_properties.pageSetUpPr.fitToPage = True  # 此行必须设置
        ws.page_setup.fitToHeight = False

        ws.print_options.gridLines = True  # 页面设置->工作表->网格线

        # 页边距
        ws.page_margins.left = 0.1  # 左
        ws.page_margins.right = 0.1  # 右
        ws.page_margins.top = 0.1  # 上
        ws.page_margins.bottom = 0.1  # 下
        ws.page_margins.header = 0.1  # 页眉
        ws.page_margins.footer = 0.1  # 页脚

        # 保存
        wb.save(io)

    def file_to_dataframe(self, io, lfile):
        pdf = pdfplumber.open(io)
        colorlist= []
        pdf_title = []
        find_idx = 0
        # 打开电子发票的PDF文件  
        for page in pdf.pages:
            # 提取文本内容 
            text = page.extract_text()
            if not text.__contains__('TRIMS') or find_idx >= 1:
                continue
            find_idx = find_idx + 1
            # print(page.extract_tables()[1])
            # 提取表格2中非None的颜色列表
            colorlist_tmp = [i for i in page.extract_tables()[1][0] if i != None and i != '']
            # 提取表格2中的数据
            colorlist = colorlist_tmp[1:]
            # 排除颜色列表中PDF含有的关键字
            not_color_key = ['SUPPLIER', 'COMPOSITION', 'DETAILS', 'TRIMS']
            color_list_title = []
            if len(colorlist) == 0:
                colorlist = [i for i in page.extract_tables()[1][1] if i != None and i != '' and i not in not_color_key]
            # 循环colorlist,用索引作为color_list_title的值
            for i in range(len(colorlist)):
                color_list_title.append('颜色' + str(i + 1))
            pdf_data = page.extract_tables()[1][1:]
            pdf_data_title = np.append(['TRIMS', '供应商', '备注', '有效幅宽/规格'], color_list_title)
            # 去掉数据尾部的空值
            for i in range(len(pdf_data)):
                pdf_data[i] = pdf_data[i][:len(pdf_data_title)]
            data_df = pd.DataFrame(data=pdf_data, columns=pdf_data_title)
            data_df.replace('\n', '', regex=True, inplace=True)
        data_df = data_df.drop(columns=['供应商', '备注', '有效幅宽/规格'])
        # 如果data_df的第一行第一列不是TRIMS, 则在第一行追加一行数据
        if data_df.iloc[0, 0] != 'TRIMS':
            first_df = pd.DataFrame([['TRIMS'] + colorlist], columns=['TRIMS'] + color_list_title)
            data_df = pd.concat([first_df, data_df], axis=0 ,ignore_index=True)
        all_data_df = pd.DataFrame([], columns=['TRIMS', '物料'])
        all_data_df['TRIMS'] = data_df['TRIMS']
        all_data_df['物料'] = data_df['颜色1']
        all_data_df['颜色'] = data_df['颜色1'][0]
        # 循环 data_df 的第三列到最后一列的内容
        for col in data_df.columns[2:]:
            tmp_data_df = pd.DataFrame([], columns=['TRIMS', '物料'])
            tmp_data_df['TRIMS'] = data_df['TRIMS']
            tmp_data_df['物料'] = data_df[col]
            tmp_data_df['颜色'] = data_df[col][0]
            # 合并all_data_df和tmp_data_df
            all_data_df = pd.concat([all_data_df, tmp_data_df], ignore_index=True)
        # 追加“款号”列
        all_data_df.insert(0, '款号', lfile)
        # 删除 TRIMS 列中包含 'SEND ME MOCK UP WITH' 和 'TRIMS' 的行
        all_data_df.fillna('', inplace=True)
        all_data_df = all_data_df[~all_data_df['TRIMS'].str.contains('SEND ME MOCK UP WITH')]
        all_data_df = all_data_df[~all_data_df['TRIMS'].str.contains('TRIMS')]
        # 重新排列索引
        all_data_df = all_data_df.reset_index(drop=True)
        if data_df.empty:
            print('没有找到TRIMS，文件名：' + lfile)
        # 导出excel
        excelUrl = self.trim_list_file_finish + '\\' + lfile + '.xlsx'
        writer = pd.ExcelWriter(excelUrl, engine='xlsxwriter')
        all_data_df.to_excel(writer, lfile, index=False)
        writer.save()

    def merge_excel_files(self):
        # 存储所有数据的列表
        all_data = []
        # 遍历指定目录及其子目录中的所有文件
        for root, dirs, files in os.walk(self.trim_list_file_finish):
            for file in files:
                if file.endswith('.xlsx') or file.endswith('.xls'):
                    file_path = os.path.join(root, file)
                    # 读取 Excel 文件
                    df = pd.read_excel(file_path)
                    # 将数据添加到列表中
                    all_data.append(df)
        # 合并所有数据
        combined_df = pd.concat(all_data, ignore_index=True)
        # print(combined_df)
        # 将合并后的数据保存到新的 Excel 文件
        excelUrl = self.trim_list_file_finish + '\\' + '合并文件.xlsx'
        with pd.ExcelWriter(excelUrl, engine='xlsxwriter') as writer:
            combined_df.to_excel(writer, '合并文件', index=False)
def gui_start():
    VAS = VAS_GUI()
    VAS.get_files()


if __name__ == '__main__':
    gui_start()
