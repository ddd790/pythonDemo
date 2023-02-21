import os
import shutil
import pandas as pd
import time
from tkinter import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Side, Border, Font
from openpyxl.worksheet.dimensions import RowDimension
from dateutil import parser


class VAS_GUI():
    # 批量获取服务器数据，进行累加操作
    def get_files(self):
        print('数据操作进行中......')
        # 追加的dataFrame的title
        self.add_data_title = ['用料名称', '品号', '颜色', '规格', '预估单耗']
        # 排除列
        self.out_column = ['FUSIBLE', 'CHEST PIECE', 'SLEEVE HEAD', 'Vendor', 'Date', 'THREAD', 'THREAD BTN HOLE LAPEL', 'WAISTBAND FUSIBLE',
                           'BARTACKS OUT', 'BARTACKS IN', 'SADDLE STITCH', 'OUTSIDE STITCH', 'INSIDE STITCH', 'COAT POCKETING',
                           'COAT POCKETING OUTSIDE', 'PANT POCKETING', 'INSIDE PANT BUTTON', 'ZIPPER', 'ZROH PANT M', 'PANT THREAD', 'VEST POCKETING', 'VEST POCKETING OUTSIDE']
        # 根据勤哲的key匹配对应trimList中的key和value
        self.local_trim_list_file = 'd:\\trimlistToBom'
        self.trim_list_file_finish = 'd:\\trimlistToBom结果'
        self.en_cn_file = r'\\192.168.0.3\01-业务一部资料\软件\trimlistToBom\en_cn.xlsx'
        self.cn_sample_file = r'\\192.168.0.3\01-业务一部资料\A-Serena\2 - 辅料表\样例'
        # 删除目录内文件
        if os.path.exists(self.trim_list_file_finish):
            shutil.rmtree(self.trim_list_file_finish)
        os.mkdir(self.trim_list_file_finish)

        # 读取中英文翻译的配置文件
        # 英文行数（默认最大是301行）
        self.max_en = 301
        self.en_cn = {}
        self.get_en_cn_config(self.en_cn_file)

        # 循环本地临时文件，处理合并
        self.table_value = []
        for lroot, ldirs, lfiles in os.walk(self.local_trim_list_file):
            for lfile in lfiles:
                print(lfile)
                self.file_to_dataframe(os.path.join(lroot, lfile), str(
                    lfile).split('.')[0])

        # 循环读取文件，修改样式
        for root, dirs, files in os.walk(self.trim_list_file_finish):
            for file in files:
                self.change_file_style(os.path.join(root, file))

        # print(self.table_value)
        print('已经完成导出操作！请到D盘【trimlistToBom结果】中查看文件吧~~~~')
        input('按回车退出 ')

    # 获取中英文对照配置文件
    def get_en_cn_config(self, io):
        # 打开工作表
        wb = load_workbook(io)
        ws = wb.active
        for i in range(1, self.max_en):
            data_en_value = ws.cell(row=i, column=1).value
            data_cn_value = ws.cell(row=i, column=2).value
            self.en_cn[data_en_value] = data_cn_value

    # 修改文件样式
    def change_file_style(self, io):
        # 打开工作表
        wb = load_workbook(io)
        ws = wb.active
        # 调整第列宽
        ws.column_dimensions['A'].width = 32.25
        ws.column_dimensions['B'].width = 46
        ws.column_dimensions['C'].width = 21.3
        ws.column_dimensions['D'].width = 55
        ws.column_dimensions['E'].width = 10

        # 定义表头颜色样式为橙色
        header_fill = PatternFill('solid', fgColor='FFFF00')
        header_fill_qing = PatternFill('solid', fgColor='00FFFF')
        header_fill_lan = PatternFill('solid', fgColor='EE82EE')

        # 定义对齐样式横向居中、纵向居中
        align = Alignment(horizontal='center',
                          vertical='center', wrapText=True)
        # 定义对齐样式纵向居中, 自动换行
        align_center = Alignment(vertical='center', wrapText=True)

        # 定义边样式为细条
        side = Side('thin')
        # 定义边框样式，有底边和右边
        border = Border(top=side, bottom=side, left=side, right=side)

        # 用料名称所在行
        row_use_name = 0
        for i in range(1, self.max_en + 5):
            data_value = ws.cell(row=i, column=1).value
            if data_value == '用料名称':
                row_use_name = i
                break

        # 设置基本数据的样式
        base_step = 0
        for row in ws.iter_rows(min_row=1, max_row=row_use_name):
            base_step = base_step + 1
            # ws.row_dimensions[base_step].height = 18
            RowDimension(ws, index=base_step, height=True)
            for cell in row:
                cell.alignment = align_center
                cell.font = Font(size=12)
                if str(cell.value).__contains__('接缝滑移'):
                    # 设置单元格填充颜色
                    cell.fill = header_fill_lan
                    cell.font = Font(bold=True)
                if str(cell.value).__contains__('面料描述'):
                    # 设置单元格填充颜色
                    cell.fill = header_fill_qing
                    # 设置单元格对齐方式
                    cell.font = Font(bold=True)

        # 设置用料名称单元格格式
        ws.row_dimensions[row_use_name].height = 16
        for cell in ws[row_use_name]:
            # 设置单元格填充颜色
            cell.fill = header_fill
            # 设置单元格对齐方式
            cell.alignment = align
            # 设置单元格边框
            cell.border = border
            cell.font = Font(bold=True)
        step = 0
        for row in ws.iter_rows(min_row=row_use_name + 1, max_row=ws.max_row):
            step = step + 1
            ws.row_dimensions[row_use_name + step].height = 16
            special_backcolor_flag = 0
            for cell in row:
                # 判断【胸兜牌衬】和【腰硬衬】，对应的D列需要背景色为黄色
                if str(cell.value).__contains__('胸兜牌衬') or str(cell.value).__contains__('腰硬衬'):
                    special_backcolor_flag = 1
                if special_backcolor_flag >= 1:
                    special_backcolor_flag = special_backcolor_flag + 1
                if special_backcolor_flag == 5:
                    # 设置单元格填充颜色
                    cell.fill = header_fill
                cell.alignment = align_center
                cell.border = border

        # 设置打印标题和打印列
        # ws.print_title_rows = '3:1'
        # ws.print_title_cols = "A:E"

        # 设置打印A4纵向
        ws.set_printer_settings(ws.PAPERSIZE_A4, ws.ORIENTATION_PORTRAIT)

        # 所有列设置为一页 逆向思维,先缩放到页面 然后适合高度改为FLASE
        ws.sheet_properties.pageSetUpPr.fitToPage = True  # 此行必须设置
        ws.page_setup.fitToHeight = False

        ws.print_options.gridLines = True  # 页面设置->工作表->网格线

        # 页边距
        ws.page_margins.left = 0.1  # 左
        ws.page_margins.right = 0.1  # 右
        ws.page_margins.top = 0.1  # 上
        ws.page_margins.bottom = 0.1  # 下
        ws.page_margins.header = 0.1  # 页眉
        ws.page_margins.footer = 0.1  # 页脚

        # 保存
        wb.save(io)

    def file_to_dataframe(self, io, lfile):
        # 读取文件
        excelData = pd.read_excel(io, header=None, keep_default_na=False)
        # csv中的title
        formartExcelTitle = []
        # csv数据
        formatExcelvalue = excelData.values
        for csvIdx in range(0, len(excelData.values[0])):
            formartExcelTitle.append(csvIdx)
        df = pd.DataFrame(formatExcelvalue, columns=formartExcelTitle)
        # title
        excelTitle = df[0]
        # 正常数据，偶数列的数据
        dataVal = []
        # 描述数据，基数列的数据
        disVal = []
        for tempIndex in formartExcelTitle:
            str_arr = df[tempIndex].values
            for arr_i in range(len(str_arr)):
                add_value = str(str_arr[arr_i]).replace(
                    '=', '').replace('"', '')
                if len(add_value) > 0:
                    add_value = self.remove_zero(add_value)
                str_arr[arr_i] = add_value
            if tempIndex != 0 and tempIndex % 2 == 0:
                disVal.append(str_arr)
            elif tempIndex % 2 != 0:
                dataVal.append(str_arr)

        valueDf = pd.DataFrame(dataVal, columns=excelTitle)
        # valueDf['version'] = version
        disDf = pd.DataFrame(disVal, columns=excelTitle)
        disDf['Purchasing Document'] = ''
        # disDf['version'] = version
        # 对excel读取的数据进行整理，整理成符合要求的格式
        self.arrange_excel_data(valueDf, disDf, lfile)

    def arrange_excel_data(self, valueDf, disDf, lfile):
        material_list = list(set(disDf['Fabric'].tolist()))
        # trimList中的key和value
        for idx, value in valueDf.iterrows():
            arrangeVal = []
            # 款号，生成文件的文件名用，文件名为po_Material_Style_idx
            # tempFileName = str(valueDf.loc[idx]['Purchasing Document']) + '_' + str(
            #     valueDf.loc[idx]['Material']) + '_' + str(idx)
            tempFileName = lfile + '_' + str(idx)
            firstflag = True
            # 相同结果的数据作为key
            sameKey = []
            sameVal = []
            sameDic = {}
            # 源文件内容的dataframe
            # oldFileData = []
            # 订单类型（上衣，裤子，马甲，套装，三件套）
            materialType = str(disDf.loc[idx]['Material'])
            # trimlist中有2种或者3中的订单类型，就是套装或者三件套
            if len(material_list) > 1:
                materialType = ','.join(material_list)
            # 标题的汉字列，填入到最后的对应列中（最终文件的B列）
            column_dic = self.get_file_name_by_num_style(
                len(list(valueDf)), materialType)
            for title in list(valueDf):
                arrangeVal.append(
                    [title, str(column_dic[title]), str(value[title]), str(disDf.loc[idx][title]), ''])
                # 从ZROH FRONT开始
                if firstflag and title.__contains__('ZROH'):
                    firstflag = False
                if firstflag == False and value[title] != '' and title not in self.out_column:
                    sameKey.append(str(value[title]) +
                                   '^_^' + str(disDf.loc[idx][title]))
                    sameVal.append(title)
            arrangeVal.append(['', '', '', '', ''])
            arrangeVal.append(['', '', '', '', ''])
            arrangeVal.append(['', '', '', '', ''])
            arrangeVal.append(['', '', '', '', ''])
            arrangeVal.append(['', '', '', '', ''])
            arrangeVal.append(['', '', '', '', ''])
            arrangeVal.append(['', '', '', '', ''])
            arrangeVal.append(['', '', '', '', ''])
            arrangeVal.append(['', '', '', '', ''])
            arrangeVal.append(['', '', '', '', ''])
            arrangeVal.append(['', '', '', '', ''])
            arrangeVal.append(['', '', '', '', ''])
            arrangeVal.append(['', '', '', '', ''])
            arrangeVal.append(['', '', '', '', ''])
            arrangeVal.append(['', '', '', '', ''])
            arrangeVal.append(self.add_data_title)

            # 相同结果的值拼接title
            for i in range(len(sameKey)):
                sameDic.setdefault(sameKey[i], []).append(
                    self.en_cn[sameVal[i]])
            # 遍历整理好的字典,进行dataframe值的整理
            for key, value in sameDic.items():
                valTempArr = []
                # 逗号分割值
                tempValString = "/".join(str(i) for i in value)
                tempkey = ''
                if tempValString.__contains__('面料') or tempValString == '补充意见 1' or tempValString == '补充意见 2' or tempValString == '补充意见 3' or tempValString == '补充意见 4':
                    tempkey = key[0:key.rfind('^_^')]
                else:
                    tempkey = key[key.rfind('^_^') + 3:]
                valTempArr.append(tempValString)
                valTempArr.append(tempkey)
                valTempArr.append('')
                # 规格如果是扣，需要追加^_^前面的
                if tempValString[-1:] == '扣':
                    kouKey = key[0:key.rfind('^_^')]
                    valTempArr.append(kouKey[kouKey.rfind('-') + 1:])
                elif tempValString == '领底绒':
                    valTempArr.append('90cm')
                else:
                    valTempArr.append('')
                valTempArr.append('')
                arrangeVal.append(valTempArr)

            basicType = str(disDf.loc[idx]['Material'])
            table_data = pd.DataFrame(self.arrange_appand_by_type(
                arrangeVal, basicType), columns=self.add_data_title)
            # 导出excel,追加在old的后面
            excelUrl = self.trim_list_file_finish + '\\' + tempFileName + '.xlsx'
            writer = pd.ExcelWriter(excelUrl, engine='xlsxwriter')
            sheet_name = str(valueDf.loc[idx]['Style']) + '_' + str(idx)
            table_data.to_excel(writer, sheet_name, index=False, header=None)
            writer.save()

    def arrange_appand_by_type(self, arrangeVal, type):
        # 根据类型追加固定字段
        if type.__contains__('coats'):
            arrangeVal.append(['兜布', 'ECO-8301 ', '本白', '146cm', ''])
            arrangeVal.append(['兜位衬', '0118N ', '白', '99cm', ''])
            arrangeVal.append(['拉丝衬', 'F0125N', '白', '99cm', ''])
            arrangeVal.append(['无胶衬', 'SF-35 ', '白', '99cm', ''])
            arrangeVal.append(['胸兜牌衬', '2346-2HE', '白', '2.1cm', ''])
            arrangeVal.append(['有纺直条', '5850-1', '白', '2.0cm', ''])
            arrangeVal.append(['端打条', '5850-3', '白', '1.2cm', ''])
            arrangeVal.append(['拉丝无纺衬条', '9332-1', '白', '1.0cm', ''])
            arrangeVal.append(['双面胶', '双面胶', '白', '0.8cm', ''])
            arrangeVal.append(['小棉带', '小棉带', '白', '0.3cm', ''])
        elif type.__contains__('Pants'):
            arrangeVal.append(['腰里上部', '', '', '', ''])
            arrangeVal.append(['腰里夹牙', '', '', '', ''])
            arrangeVal.append(['裤口袋布', '涤棉人字纹-ECO-4303P', '黑', '146cm', ''])
            arrangeVal.append(['前门襟拉链', 'CFC-36 DA3', '', '', ''])
            arrangeVal.append(['裤钩', 'B498 ', '亮银色', '', ''])
            arrangeVal.append(['裤内扣', 'SB', '黑', '22L', ''])
            arrangeVal.append(['无纺衬-小部位', 'PE125 ', '炭灰', '150cm', ''])
            arrangeVal.append(['无纺衬-腰里下部', 'PE125 ', '炭灰', '150cm', ''])
            arrangeVal.append(['腰硬衬', 'FW6951M68 ', '黑', '3.3cm', ''])
            arrangeVal.append(['绊带衬', '4947', '黑', '0.9cm', ''])
            arrangeVal.append(['腰网衬', '6148', '黑', '5.5cm', ''])
            arrangeVal.append(['双面胶', '双面胶', '白', '0.8cm', ''])
            arrangeVal.append(['PL腰里加工', 'PL腰里加工', '', '', ''])
        elif type.__contains__('Vests'):
            arrangeVal.append(['马甲钎子', 'BG87-006JZ', '古铜色', '', ''])
            arrangeVal.append(['兜布', 'ECO-8301', '黑', '146cm', ''])
            arrangeVal.append(['前身衬', 'PE206 ', '黑', '148cm', ''])
            arrangeVal.append(['腰兜牌衬', '2346-2HE', '白', '', ''])
            arrangeVal.append(['无纺衬（小部位）', 'PE125', '炭灰', '150cm', ''])
            arrangeVal.append(['拉丝衬条', '9332-1', '黑', '1cm', ''])
            arrangeVal.append(['双面胶', '双面胶', '白', '0.8cm', ''])
        elif type.__contains__('2 Piece Suits'):
            arrangeVal.append(['上衣口袋布', 'ECO-8301', '黑', '146cm', ''])
            arrangeVal.append(['兜位衬', '0118N ', '黑', '99cm', ''])
            arrangeVal.append(['拉丝衬', 'F0125N ', '黑', '99cm', ''])
            arrangeVal.append(['无胶衬', 'SF-35 ', '黑', '99cm', ''])
            arrangeVal.append(['胸兜牌衬', '2346-2HE ', '白', '2.1cm', ''])
            arrangeVal.append(['有纺直条', '5850-1 ', '黑', '2.0cm', ''])
            arrangeVal.append(['拉丝无纺衬条', '9332-1 ', '黑', '1.0cm', ''])
            arrangeVal.append(['双面胶 上衣+裤子', '双面胶 ', '白', '0.8cm', ''])
            arrangeVal.append(['小棉带', '小棉带 ', '黑', '0.3cm', ''])
            arrangeVal.append(['端打条', '5850-3 ', '黑', '1.2cm', ''])
            arrangeVal.append(['裤口袋布', '涤棉人字纹-ECO-4303P', '黑', '146cm', ''])
            arrangeVal.append(['前门襟拉链', 'CFC-36 DA3', '', '', ''])
            arrangeVal.append(['裤钩', 'B498 ', '亮银色', '', ''])
            arrangeVal.append(['裤内扣', 'SB', '黑', '22L', ''])
            arrangeVal.append(['腰硬衬', 'FW6951M68 ', '黑', '3.3cm', ''])
            arrangeVal.append(['绊带衬', '4947', '黑', '0.9cm', ''])
            arrangeVal.append(['腰网衬', '6148', '黑', '5.5cm', ''])
            arrangeVal.append(['PL腰里加工', 'PL腰里加工', '', '', ''])
        else:
            arrangeVal.append(['口袋布 上衣+马甲', 'ECO-8301', '黑', '146cm', ''])
            arrangeVal.append(['兜位衬', '0118N ', '黑', '99cm', ''])
            arrangeVal.append(['拉丝衬', 'F0125N ', '黑', '99cm', ''])
            arrangeVal.append(['无胶衬', 'SF-35 ', '黑', '99cm', ''])
            arrangeVal.append(['胸兜牌衬 上衣+马甲', '2346-2HE ', '白', '2.1cm', ''])
            arrangeVal.append(['有纺直条', '5850-1 ', '黑', '2.0cm', ''])
            arrangeVal.append(['拉丝无纺衬条 上衣+马甲', '9332-1 ', '黑', '1.0cm', ''])
            arrangeVal.append(['双面胶 上衣+裤子+马甲', '双面胶 ', '白', '0.8cm', ''])
            arrangeVal.append(['小棉带', '小棉带 ', '黑', '0.3cm', ''])
            arrangeVal.append(['端打条', '5850-3 ', '黑', '1.2cm', ''])
            arrangeVal.append(['裤口袋布', '涤棉人字纹-ECO-4303P', '黑', '146cm', ''])
            arrangeVal.append(['前门襟拉链', 'CFC-36 DA3', '', '', ''])
            arrangeVal.append(['裤钩', 'B498', '亮银色', '', '1.00'])
            arrangeVal.append(['裤内扣', 'SB', '黑', '22L', ''])
            arrangeVal.append(['腰硬衬', 'FW6951M68 ', '黑', '3.3cm', ''])
            arrangeVal.append(['绊带衬', '4947', '黑', '0.9cm', ''])
            arrangeVal.append(['腰网衬', '6148', '黑', '5.5cm', ''])
            arrangeVal.append(['马甲钎子', 'BG87-006JZ', '古铜色', '', ''])
            arrangeVal.append(['PL腰里加工', 'PL腰里加工', '', '', ''])
        return arrangeVal

    def get_file_name_by_num_style(self, num, style):
        # 根据类型和数字返回对应的文件名
        file_content_name = ''
        if (style.__contains__('coats') and style.__contains__('Pants') and style.__contains__('Vests')) or (style.__contains__('3 Piece Suits')):
            file_content_name = '三'
        elif (style.__contains__('coats') and style.__contains__('Pants')) or (style.__contains__('coats') and style.__contains__('Vests')):
            file_content_name = '套'
        elif style.__contains__('Pants') and style.__contains__('Vests'):
            file_content_name = '套'
        elif style.__contains__('coats'):
            file_content_name = '上衣'
        elif style.__contains__('Pants'):
            file_content_name = '裤子'
        elif style.__contains__('Vests'):
            file_content_name = '马'
        # elif style.__contains__('3 Piece Suits'):
        #     file_content_name = '三'
        else:
            file_content_name = '套'
        last_file_name = ''
        # 基础时间，设置为2018年9月
        last_time = parser.parse('2018 9 03')
        for root, dirs, files in os.walk(self.cn_sample_file):
            for file in files:
                # 修改时间
                mtime = parser.parse(time.ctime(os.path.getmtime(
                    os.path.join(root, file))))
                diff_time = mtime - last_time
                if file_content_name == '套':
                    if str(file).__contains__(file_content_name) and str(file).__contains__(str(num)) and not str(file).__contains__('三') and diff_time.days > 0:
                        last_time = mtime
                        last_file_name = file
                    elif str(file).__contains__(file_content_name) and str(file).__contains__(str(num)) and (str(file).__contains__('假') or str(file).__contains__('两')) and diff_time.days > 0:
                        last_time = mtime
                        last_file_name = file
                else:
                    if str(file).__contains__(file_content_name) and str(file).__contains__(str(num)) and diff_time.days > 0:
                        last_time = mtime
                        last_file_name = file
        # 读取文件版本最大的
        io = self.cn_sample_file + '\\' + last_file_name
        cn_column_dic = {}
        excelKey = pd.read_excel(
            io, header=None, keep_default_na=False, usecols=[0])
        excelValue = pd.read_excel(
            io, header=None, keep_default_na=False, usecols=[1])
        # 第一列结果赋值list
        key_result = []
        for s_li in excelKey.values.tolist():
            key_result.append(s_li[0])
        # 第二列结果赋值list
        value_result = []
        for v_li in excelValue.values.tolist():
            value_result.append(v_li[0])
        # 转为字典
        cn_column_dic = dict(zip(key_result, value_result))
        return cn_column_dic

    def remove_zero(self, str):
        while str[0] == "0":
            str = str[1:]
        return str


def gui_start():
    VAS = VAS_GUI()
    VAS.get_files()


if __name__ == '__main__':
    gui_start()
