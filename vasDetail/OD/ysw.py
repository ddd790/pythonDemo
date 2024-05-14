import os
import win32com.client
import gc
import pandas as pd
import datetime
import shutil
from docxtpl import DocxTemplate
import docx
from docxcompose.composer import Composer
from glob import glob
import time
from PyPDF2 import PdfFileReader, PdfFileWriter
import openpyxl


class VAS_GUI():
    # 批量获取服务器数据，进行累加操作
    def get_files(self):
        print('数据操作进行中......' + str(datetime.datetime.now()).split('.')[0])
        self.local_trim_list_file = 'd:\\亚瑟王花名册'
        self.trim_list_file_finish = 'd:\\亚瑟王模板'
        # 删除目录内文件
        # if os.path.exists(self.trim_list_file_finish):
        #     shutil.rmtree(self.trim_list_file_finish)
        # os.mkdir(self.trim_list_file_finish)
        # 新建 pdf 文件夹，所有生成的 PDF 文件都放在里面
        folder = self.trim_list_file_finish + '\\pdf\\'
        if os.path.exists(folder):
             shutil.rmtree(folder)
        os.makedirs(folder)
        # 新建 excel 文件夹，所有生成的 excel 文件都放在里面
        folder = self.trim_list_file_finish + '\\excel\\'
        if os.path.exists(folder):
             shutil.rmtree(folder)
        os.makedirs(folder)

        # 合并excel的文件
        # all_file_path = self.trim_list_file_finish + '\\0all.xlsx'
        # if os.path.exists(all_file_path):
        #     os.remove(all_file_path)

        # 循环文件,拆分成各自模板
        print('开始生成EXCEL..........')
        for lroot, ldirs, lfiles in os.walk(self.local_trim_list_file):
            for lfile in lfiles:
                if str(lfile).__contains__('.xls') or str(lfile).__contains__('.xlsx'):
                    self.file_to_dataframe(os.path.join(lroot, lfile))
        # excel转PDF
        # 将目标文件夹所有文件归类，转换时只打开一个进程
        xls = []
        for fn in os.listdir(self.trim_list_file_finish + '\\excel\\'):
            if fn.endswith(('.xls', '.xlsx')):
                xls.append(fn)

        # 新建 pdf 文件夹，所有生成的 PDF 文件都放在里面。
        self.EXCEL2pdf(self.trim_list_file_finish, xls)

        # 合并所有pdf文件为一个pdf
        print('开始合并PDF..........')
        all_file_path = self.trim_list_file_finish + '\\0all.pdf'
        if os.path.exists(all_file_path):
            os.remove(all_file_path)
        file_writer = PdfFileWriter()
        for root, dirs, files in os.walk(self.trim_list_file_finish + '\\pdf'):
            for file in files:
                if str(file).__contains__('.pdf'):
                    # 循环读取需要合并pdf文件
                    file_reader = PdfFileReader(os.path.join(root, file))
                    # 遍历每个pdf的每一页
                    for page in range(file_reader.getNumPages()):
                        # 写入实例化对象中
                        file_writer.addPage(file_reader.getPage(page))
        with open(self.trim_list_file_finish + '\\0all.pdf', 'wb') as out:
            file_writer.write(out)

        print('已经完成操作！' + str(datetime.datetime.now()).split('.')[0])
        input('按回车退出 ')

    def file_to_dataframe(self, io):
        excel_header = ['序号', '姓名', '所在部门', '岗位', '入职时间', '出生日期', '身份证号', '地址', '性别', '政治面貌', '学历', '人员类别', '手机']
        excelData = pd.read_excel(io, header=2, keep_default_na=False)
        df = pd.DataFrame(excelData.values, columns=excel_header)
        # 循环df内容，进行赋值
        read_doc_file_path = self.trim_list_file_finish + '\\亚瑟王员工信息表.xlsx'
        all_excel_file = self.trim_list_file_finish + '\\0all.xlsx'
        for i in range(len(df)):
            data=openpyxl.load_workbook(read_doc_file_path)
            sheet_name = str(df.iloc[i]['序号']) + str(df.iloc[i]['姓名'])
            # 导出excel
            excelUrl = self.trim_list_file_finish + '\\excel\\' + sheet_name + '.xlsx'

            sheet = data.active
            sheet.title = sheet_name
            sheet['B3'] = str(df.iloc[i]['姓名'])
            sheet['F3'] = str(df.iloc[i]['性别'])
            sheet['H3'] = str(df.iloc[i]['政治面貌'])
            sheet['D5'] = str(df.iloc[i]['学历'])
            sheet['F5'] = str(df.iloc[i]['出生日期'])[0:10]
            sheet['B7'] = str(df.iloc[i]['地址'])
            sheet['C9'] = str(df.iloc[i]['身份证号'])
            sheet['I11'] = str(df.iloc[i]['手机'])
            sheet['C15'] = str(df.iloc[i]['入职时间'])[0:10]
            sheet['F15'] = str(df.iloc[i]['所在部门'])
            sheet['I15'] = str(df.iloc[i]['岗位'])
            data.save(excelUrl)
            # if os.path.exists(all_excel_file):
            #     data_all=openpyxl.load_workbook(all_excel_file)
            #     data_all.create_sheet(sheet_name)
            #     append_data = list(data[sheet_name].values)
            #     for i in range(0, len(append_data)):
            #         data_all[sheet_name].append(append_data[i])
            #     data_all.save(all_excel_file)
            # else:
            #     data.save(all_excel_file)


    # EXCEL2pdf
    def EXCEL2pdf(self, filePath, xls):
        # 如果没有文件则提示后直接退出
        if (len(xls) < 1):
            print("\n【无 EXCEL 文件】\n")
            return
        xlApp = win32com.client.DispatchEx("Excel.Application")
        xlApp.Visible = False
        xlApp.DisplayAlerts = 0
        # 创建一个空的exce作为合并用
        # writer = openpyxl.Workbook()
        try:
            # 开始转换
            print("\n【开始 EXCEL -> PDF 转换】")
            for i in range(len(xls)):
                fileName = xls[i]  # 文件名称
                fromFile = os.path.join(filePath + '\\excel\\', fileName)  # 文件地址
                toFile = self.changeSufix2Pdf(os.path.join(filePath + '\\pdf\\', fileName))  # 生成的文件地址
                try:
                    # 某文件出错不影响其他文件打印
                    # time.sleep(1)
                    books = xlApp.Workbooks.Open(fromFile,False)
                    books.ExportAsFixedFormat(0, toFile)  # 生成的所有 PDF 都会在 PDF 文件夹中
                    # # 读取excel文件
                    # df = openpyxl.load_workbook(fromFile)
                    # # 获取文件名作为sheet名
                    # sheet_name = df.worksheets[0]
                    # # 将数据写到新的excel文件的不同sheet页中
                    # df.to_excel(writer, sheet_name = sheet_name, index=False)
                    print("转换到：【"+toFile+"】完成")
                    # 关闭 excel 进程
                    books.Close(False)
                except Exception as e:
                    print(e)
            xlApp.Quit()
        except Exception as e:
            print(e)
        finally:
            gc.collect()

    # 修改后缀名
    def changeSufix2Pdf(self, file):
        return file[:file.rfind('.')]+".pdf"

    # 合并excel文档
    # def merge_file(self, files_list):
    #     workbook = openpyxl.Workbook()
    #     for lroot, ldirs, lfiles in os.walk(self.trim_list_file_finish + '\\excel'):
    #         for lfile in lfiles:
    #             if str(lfile).__contains__('.xls') or str(lfile).__contains__('.xlsx'):
    #                 writer=openpyxl.load_workbook(os.path.join(lroot, lfile))
    #                 workbook.create_sheet(str(lfile).split('.xlsx')[0])

    #     workbook.save(os.path.join(self.trim_list_file_finish + '\\0all.xlsx'))

def gui_start():
    VAS = VAS_GUI()
    VAS.get_files()


if __name__ == '__main__':
    gui_start()
